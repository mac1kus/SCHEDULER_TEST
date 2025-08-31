from flask import render_template, request, jsonify, send_file, redirect, url_for
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.series import Series
import tempfile
import json
from collections import defaultdict
from dotenv import load_dotenv

load_dotenv()
APP_USERNAME = os.environ.get("APP_USERNAME")
APP_PASSWORD = os.environ.get("APP_PASSWORD")


from utils import (
    AdvancedRefineryCrudeScheduler,
    get_date_with_ordinal,
    _parse_json_datetime,
    _save_excel_with_conflict_handling,
    _calculate_timestamp_consumption_summary,
    populate_tank_times
)

# Global scheduler instance
scheduler = AdvancedRefineryCrudeScheduler()

# Save/Load user inputs configuration
INPUTS_FILE = "last_inputs.json"

def register_routes(app):
    """Register all routes with the Flask app"""

    @app.route('/login')
    def login_page():
         message = request.args.get('message', '')
         return render_template('login.html', message=message)

    @app.route('/')
    def root():
        return redirect('/login')


    @app.route('/login', methods=['POST'])
    def login():
        username = request.form.get('username')
        password = request.form.get('password')
        if username == APP_USERNAME and password == APP_PASSWORD:
           return redirect('/index')
        else:
            return redirect(url_for('login_page', message='Invalid username or password'))
    
    @app.route('/index')
    def index():
        return render_template('index.html')

    @app.route('/api/simulate', methods=['POST'])
    def simulate():
        params = request.json
        results = scheduler.run_simulation(params)
        return jsonify(results)

    # THE CORRECTED VERSION
    def _create_sequence_summary_sheets(wb, results):
        """Create ONE sheet with all 3 sequence tables"""
        try:
            ws = wb.create_sheet("Sequence Summary")

            # 1. Create the timestamp string
            timestamp_str = f"Report Generated On: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}"
            
            # 2. THIS IS THE MISSING STEP: Write the string to cell A1
            timestamp_cell = ws.cell(row=1, column=1, value=timestamp_str)
            timestamp_cell.font = Font(bold=True, italic=True, color="4F4F4F")
            
            # Extract data from simulation results
            cargo_report = results.get('cargo_report', [])
            feeding_events_log = results.get('feeding_events_log', [])
            filling_events_log = results.get('filling_events_log', [])
            
            # 3. Start the main report from row 3 to make space for the timestamp
            current_row = 3
            
            # 4. CARGO SEQUENCE TABLE (now starts on row 3)
            ws.cell(row=current_row, column=1, value="CARGO SEQUENCE").font = Font(bold=True, size=14)
            current_row += 2
            
            # (the rest of your function continues as before)
            cargo_headers = ['CARGO', 'ARRIVAL_DATE', 'ARRIVAL_TIME', 'DEPARTURE_DATE', 'DEPARTURE_TIME']
            for col, header in enumerate(cargo_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            current_row += 1
            
            for cargo in cargo_report:
                arrival_time = cargo.get('arrival_time', '')
                departure_time = cargo.get('dep_unload_port', '')
                
                arrival_date, arrival_time_only = '', ''
                if arrival_time and isinstance(arrival_time, str) and '/' in arrival_time:
                    parts = arrival_time.split(' ')
                    arrival_date = parts[0]
                    arrival_time_only = parts[1] if len(parts) > 1 else ''
                
                departure_date, departure_time_only = '', ''
                if departure_time and isinstance(departure_time, str) and '/' in departure_time:
                    parts = departure_time.split(' ')
                    departure_date = parts[0]
                    departure_time_only = parts[1] if len(parts) > 1 else ''
                
                row_data = [
                    cargo.get('type', '').title(),
                    arrival_date, arrival_time_only,
                    departure_date, departure_time_only
                ]
                
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=current_row, column=col, value=value).alignment = Alignment(horizontal='center')
                current_row += 1
            
            current_row += 2
            
            # 2. FEEDING SEQUENCE TABLE
            ws.cell(row=current_row, column=1, value="FEEDING SEQUENCE").font = Font(bold=True, size=14)
            current_row += 2
            
            feeding_headers = ['TANK', 'START_DATE', 'START_TIME', 'END_DATE', 'END_TIME']
            for col, header in enumerate(feeding_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            current_row += 1
            
            sorted_feeding_log = sorted(feeding_events_log, key=lambda x: _parse_json_datetime(x.get('start')) if x.get('start') else datetime.min)

            for event in sorted_feeding_log:
                start_dt = _parse_json_datetime(event.get('start'))
                end_dt = _parse_json_datetime(event.get('end'))
                
                row_data = [
                    f"Tank {event.get('tank_id', '')}",
                    start_dt.strftime('%d/%m/%y') if start_dt else '',
                    start_dt.strftime('%H:%M') if start_dt else '',
                    end_dt.strftime('%d/%m/%y') if end_dt else 'N/A',
                    end_dt.strftime('%H:%M') if end_dt else 'N/A'
                ]
                
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=current_row, column=col, value=value).alignment = Alignment(horizontal='center')
                current_row += 1
            
            current_row += 2
            
            # 3. FILLING, SETTLING & LAB TESTING SEQUENCE TABLE
            ws.cell(row=current_row, column=1, value="FILLING, SETTLING & LAB TESTING SEQUENCE").font = Font(bold=True, size=14)
            current_row += 2

            processing_headers = ['TANK', 'FILL_START_DATE', 'FILL_START_TIME', 'FILL_END_DATE', 'FILL_END_TIME', 'SETTLE_START_DATE', 'SETTLE_START_TIME', 'LABTEST_START_DATE', 'LABTEST_START_TIME', 'READY_DATE', 'READY_TIME']
            for col, header in enumerate(processing_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            current_row += 1

            sorted_filling_events = sorted(filling_events_log, key=lambda x: _parse_json_datetime(x.get('start')) if x.get('start') else datetime.min)
            
            processed_rows = set()

            for event in sorted_filling_events:
                fill_start = _parse_json_datetime(event.get('start'))
                fill_end = _parse_json_datetime(event.get('end'))
                settle_start = _parse_json_datetime(event.get('settle_start'))
                lab_start = _parse_json_datetime(event.get('lab_start'))
                ready_time = _parse_json_datetime(event.get('ready_time'))

                row_identifier = (
                    event.get('tank_id'),
                    fill_start.strftime('%Y-%m-%d %H:%M') if fill_start else None,
                    fill_end.strftime('%Y-%m-%d %H:%M') if fill_end else None,
                    settle_start.strftime('%Y-%m-%d %H:%M') if settle_start else None,
                    lab_start.strftime('%Y-%m-%d %H:%M') if lab_start else None,
                    ready_time.strftime('%Y-%m-%d %H:%M') if ready_time else None
                )
                
                if row_identifier not in processed_rows:
                    processed_rows.add(row_identifier)
                    
                    row_data = [
                        f"Tank {event.get('tank_id', '')}",
                        fill_start.strftime('%d/%m/%y') if fill_start else '',
                        fill_start.strftime('%H:%M') if fill_start else '',
                        fill_end.strftime('%d/%m/%y') if fill_end else '',
                        fill_end.strftime('%H:%M') if fill_end else '',
                        settle_start.strftime('%d/%m/%y') if settle_start else '',
                        settle_start.strftime('%H:%M') if settle_start else '',
                        lab_start.strftime('%d/%m/%y') if lab_start else '',
                        lab_start.strftime('%H:%M') if lab_start else '',
                        ready_time.strftime('%d/%m/%y') if ready_time else '',
                        ready_time.strftime('%H:%M') if ready_time else ''
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        ws.cell(row=current_row, column=col, value=value).alignment = Alignment(horizontal='center')
                    current_row += 1

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
            
            return True
            
        except Exception as e:
            print(f"Error creating sequence summary: {str(e)}")
            return False

    def _create_tank_filling_volumes_sheet(wb, results):
        """
        Create the Tank Filling Volumes sheet with subtotals for each distinct filling operation.
        """
        try:
            ws = wb.create_sheet("Tank Filling Volumes")
            
            daily_discharge_log = results.get('daily_discharge_log', [])
            
            current_row = 1
            ws.cell(row=current_row, column=1, value="DAILY CARGO DISCHARGE").font = Font(bold=True, size=14)
            current_row += 2
            
            headers = ['DATE', 'CARGO', 'DISCHARGE (bbls)', 'TANK', 'VOL_FILLED (bbls)']
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=current_row, column=col, value=h)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                c.alignment = Alignment(horizontal='center')
            current_row += 1
        
            # Step 1: Consolidate data per day
            consolidated_data = {}
            for entry in daily_discharge_log:
                key = (entry['date'], entry['cargo_type'], entry['tank_id'])
                if key not in consolidated_data:
                    consolidated_data[key] = 0
                consolidated_data[key] += entry['volume_filled']
                
            report_events = []
            for (date, cargo_type, tank_id), volume in consolidated_data.items():
                report_events.append({
                    'date': date,
                    'cargo_type': cargo_type,
                    'tank_id': tank_id,
                    'volume_filled': volume
                })
                
            # Step 2: Sort events strictly chronologically by date
            def get_sort_datetime(event):
                try:
                    return datetime.strptime(event['date'], '%d/%m/%y')
                except (ValueError, TypeError):
                    return datetime.min
            
            report_events.sort(key=get_sort_datetime)
            
            # Step 3: Write events to sheet, adding subtotals for each filling operation
            if not report_events:
                return True # Handle case with no filling events

            # Use an iterator to allow looking ahead
            events_iterator = iter(report_events)
            current_event = next(events_iterator, None)
            
            operation_subtotal = 0
            
            while current_event:
                tank_id = current_event['tank_id']
                operation_subtotal += current_event['volume_filled']
                
                # Write the current event's data
                ws.cell(row=current_row, column=1, value=current_event['date']).alignment = Alignment(horizontal='center')
                ws.cell(row=current_row, column=2, value=current_event['cargo_type']).alignment = Alignment(horizontal='center')
                
                # *** MODIFICATION: Write as number and apply format ***
                discharge_cell = ws.cell(row=current_row, column=3, value=current_event['volume_filled'])
                discharge_cell.number_format = '#,##0'
                discharge_cell.alignment = Alignment(horizontal='center')

                ws.cell(row=current_row, column=4, value=f"Tank {current_event['tank_id']}").alignment = Alignment(horizontal='center')

                # *** MODIFICATION: Write as number and apply format ***
                vol_filled_cell = ws.cell(row=current_row, column=5, value=operation_subtotal)
                vol_filled_cell.number_format = '#,##0'
                vol_filled_cell.alignment = Alignment(horizontal='center')
                
                current_row += 1

                # Look at the next event to decide if the operation has ended
                next_event = next(events_iterator, None)
                
                operation_ended = False
                if next_event is None:
                    # This was the last event overall
                    operation_ended = True
                else:
                    # Check if the tank changes or if the dates are not consecutive
                    current_date = get_sort_datetime(current_event)
                    next_date = get_sort_datetime(next_event)
                    if next_event['tank_id'] != tank_id or (next_date - current_date).days > 1:
                        operation_ended = True

                if operation_ended:
                    # Print the subtotal for the completed operation
                    subtotal_cell = ws.cell(row=current_row, column=4, value=f"Subtotal Tank {tank_id}")
                    subtotal_cell.font = Font(bold=True)
                    subtotal_cell.alignment = Alignment(horizontal='right')
                    
                    # *** MODIFICATION: Write as number and apply format ***
                    volume_cell = ws.cell(row=current_row, column=5, value=operation_subtotal)
                    volume_cell.font = Font(bold=True)
                    volume_cell.number_format = '#,##0'
                    volume_cell.alignment = Alignment(horizontal='center')
                    current_row += 1
                    operation_subtotal = 0 # Reset for the next operation
                
                # Move to the next event
                current_event = next_event

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add a border around all the cells
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            for row in ws.iter_rows(min_row=1, max_row=current_row-1):
                for cell in row:
                    if cell.value is not None:
                        cell.border = thin_border
            
            return True
            
        except Exception as e:
            print(f"Error creating daily discharge sheet with correct subtotals: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def _create_daily_tank_status_sheet(wb, results):
        try:
            ws = wb.create_sheet("Daily Tank Status")

            # --- Color and Priority Mapping for Each Status ---
            status_styles = {
                'SUSPENDED': {'fill': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), 'priority': 1},
                'READY':     {'fill': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"), 'priority': 2},
                'LAB TEST':  {'fill': PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"), 'priority': 3},
                'SETTLING':  {'fill': PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"), 'priority': 4},
                'FILLED':    {'fill': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"), 'priority': 5},
                'FILLING':   {'fill': PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"), 'priority': 6},
                'FEEDING':   {'fill': PatternFill(start_color="F5DEB3", end_color="F5DEB3", fill_type="solid"), 'priority': 7},
                'EMPTY':     {'fill': PatternFill(fill_type=None), 'priority': 8}
            }

            # --- Data Aggregation ---
            from collections import defaultdict
            events_by_date = defaultdict(list)
            ready_counts_by_date = defaultdict(int)
            
            # Pull data from the simulated results object
            filling_events_log = results.get('filling_events_log', [])
            feeding_events_log = results.get('feeding_events_log', [])
            tank_status_log = results.get('tank_status_log', [])
            simulation_data = results.get('simulation_data', [])
            
            # Get the crudeProcessingDate to determine the year
            crude_processing_date = results.get('parameters', {}).get('crudeProcessingDate', '')
            base_year = 2024  # Default year
            if crude_processing_date:
                try:
                    # Parse the crudeProcessingDate to get the year
                    if ' ' in crude_processing_date:
                        date_part = crude_processing_date.split(' ')[0]
                        if '-' in date_part:
                            base_year = int(date_part.split('-')[0])
                        elif '/' in date_part:
                            parts = date_part.split('/')
                            if len(parts[2]) == 4:
                                base_year = int(parts[2])
                            else:
                                base_year = 2000 + int(parts[2])
                except:
                    pass
            
            # Pull initial tank levels from the parameters
            initial_stock_by_tank = {int(k.replace('tank', '').replace('Level', '')): v for k, v in results.get('parameters', {}).items() if 'tank' in k and 'Level' in k}
            
            all_events = []
            # Process all event types and collect them in a single list
            for event in filling_events_log:
                all_events.append({'dt': _parse_json_datetime(event.get('start')), 'tank_id': event.get('tank_id'), 'status': 'FILLING', 'stock': event.get('start_level', 0)})
                all_events.append({'dt': _parse_json_datetime(event.get('end')), 'tank_id': event.get('tank_id'), 'status': 'FILLED', 'stock': event.get('filled_volume', 0)})
                all_events.append({'dt': _parse_json_datetime(event.get('settle_start')), 'tank_id': event.get('tank_id'), 'status': 'SETTLING', 'stock': event.get('filled_volume', 0)})
                all_events.append({'dt': _parse_json_datetime(event.get('lab_start')), 'tank_id': event.get('tank_id'), 'status': 'LAB TEST', 'stock': event.get('filled_volume', 0)})
                all_events.append({'dt': _parse_json_datetime(event.get('ready_time')), 'tank_id': event.get('tank_id'), 'status': 'READY', 'stock': event.get('filled_volume', 0)})
            for event in feeding_events_log:
                all_events.append({'dt': _parse_json_datetime(event.get('start')), 'tank_id': event.get('tank_id'), 'status': 'FEEDING', 'stock': event.get('start_level', 0)})
                all_events.append({'dt': _parse_json_datetime(event.get('end')), 'tank_id': event.get('tank_id'), 'status': 'EMPTY', 'stock': 0})
            for entry in tank_status_log:
                if entry.get('status') == 'suspended':
                    all_events.append({'dt': _parse_json_datetime(entry.get('timestamp')), 'tank_id': entry.get('tank_id'), 'status': 'SUSPENDED', 'stock': entry.get('current_level', 0)})
            
            # Add initial states for all tanks
            for tank_id, level in initial_stock_by_tank.items():
                all_events.append({'dt': _parse_json_datetime(results.get('initial_start_time')), 'tank_id': tank_id, 'status': 'INITIAL', 'stock': level})

            # Filter out invalid events
            valid_events = [e for e in all_events if e['dt'] and e['tank_id']]
            
            # Sort all events chronologically
            valid_events.sort(key=lambda x: (x['tank_id'], x['dt']))

            # --- Add 'SUSPENDED' events based on your specific rule ---
            processed_events = []
            for i, event in enumerate(valid_events):
                processed_events.append(event)
                # Check for the specific pattern: FILLING -> ... -> FILLING -> FILLED
                if event['status'] == 'FILLING' and i + 2 < len(valid_events):
                    if valid_events[i + 1]['status'] == 'FILLING' and valid_events[i + 1]['tank_id'] == event['tank_id']:
                        if valid_events[i + 2]['status'] == 'FILLED' and valid_events[i + 2]['tank_id'] == event['tank_id']:
                            # The timestamps must be separated by a gap of more than 41 hours
                            time_difference = (valid_events[i + 1]['dt'] - event['dt']).total_seconds() / 3600
                            if time_difference > 41:
                                # Add a suspended event in the gap
                                processed_events.append({
                                    'dt': event['dt'] + timedelta(seconds=1),
                                    'tank_id': event['tank_id'],
                                    'status': 'SUSPENDED',
                                    'stock': event['stock']
                                })

            # Sort again to make sure everything is in order
            processed_events.sort(key=lambda x: (x['tank_id'], x['dt']))
            
            # Group valid events by date - now keeping track of the full date
            events_by_full_date = defaultdict(list)
            for event in processed_events:
                if event['dt'] and event['tank_id']:
                    # Store with full date for display and lookup
                    full_date = event['dt'].date()
                    events_by_full_date[full_date].append(event)
                    if event['status'] == 'READY':
                        date_key = event['dt'].strftime('%d-%b')
                        ready_counts_by_date[date_key] += 1

            if not events_by_full_date:
                ws.cell(row=1, column=1, value="No tank events recorded.")
                return True

            # --- Create Headers ---
            current_row = 1
            num_tanks = int(results.get('parameters', {}).get('numTanks', 12))
            headers = ['DAY'] + [f'TK{i}' for i in range(1, num_tanks + 1)] + ['READY'] + ['STOCKS']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            current_row += 1

            # --- Populate Data ---
            # Create a map for quick access to daily simulation records
            records_by_date = {}
            for d in simulation_data:
                ds = d.get('date')
                if ds:
                    try:
                        # Parse the date and store with date object as key
                        parsed_date = datetime.strptime(ds, '%d/%m/%y').date()
                        records_by_date[parsed_date] = d
                    except Exception as e:
                        print(f"Error parsing simulation date {ds}: {e}")
                        pass

            # Sort events by full date
            sorted_dates = sorted(events_by_full_date.keys())
            
            for event_date in sorted_dates:
                daily_events = events_by_full_date[event_date]
                daily_events.sort(key=lambda x: (x['tank_id'], x['dt']))
                
                # Format date for display
                date_str = event_date.strftime('%d-%b')
                
                start_row_for_date = current_row
                
                # Look up the simulation data for this specific date
                day_record = records_by_date.get(event_date)
                
                # Get the total stock for the day
                daily_total_stock = 0
                if day_record:
                    daily_total_stock = day_record.get('start_inventory', 0)
                
                for event in daily_events:
                    tank_id = event['tank_id']
                    status = event['status']

                    # ✅ FIXED: Use appropriate stock value based on status type
                    if status in ['EMPTY', 'FILLING', 'FILLED', 'SETTLING', 'LAB TEST', 'READY', 'SUSPENDED']:
                        # For volume-based statuses, show current tank volume (closing_stock)
                        stock_value = day_record.get(f"tank{tank_id}_closing_stock", 0) if day_record else 0
                    else:
                        # For consumption-based statuses (FEEDING), show opening stock
                        stock_value = day_record.get(f"tank{tank_id}_opening_stock", 0) if day_record else 0
                    
                    # Format cell with status, stock, and time for ALL statuses
                    time_value = event['dt'].strftime('%H:%M') if event['dt'] else 'N/A'
                    cell_value = f"STATUS: {status}\nSTOCK: {stock_value:,.0f}\nTIME: {time_value}"

                    cell = ws.cell(row=current_row, column=tank_id + 1, value=cell_value)
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)
                    
                    # Apply color based on status
                    if status in status_styles:
                        cell.fill = status_styles[status]['fill']
                    
                    current_row += 1

                end_row_for_date = current_row - 1
                
                # Write and merge date cell
                date_cell = ws.cell(row=start_row_for_date, column=1, value=date_str)
                date_cell.alignment = Alignment(horizontal='center', vertical='center')
                if start_row_for_date < end_row_for_date:
                    ws.merge_cells(start_row=start_row_for_date, start_column=1, end_row=end_row_for_date, end_column=1)

                # Write and merge READY count cell
                ready_count = ready_counts_by_date.get(date_str, 0)
                ready_cell = ws.cell(row=start_row_for_date, column=len(headers) - 1, value=ready_count)
                ready_cell.alignment = Alignment(horizontal='center', vertical='center')
                if start_row_for_date < end_row_for_date:
                    ws.merge_cells(start_row=start_row_for_date, start_column=len(headers) - 1, end_row=end_row_for_date, end_column=len(headers) - 1)

                # Write and merge total STOCKS cell
                stock_cell = ws.cell(row=start_row_for_date, column=len(headers), value=daily_total_stock)
                stock_cell.alignment = Alignment(horizontal='center', vertical='center')
                stock_cell.number_format = '#,##0'
                if start_row_for_date < end_row_for_date:
                    ws.merge_cells(start_row=start_row_for_date, start_column=len(headers), end_row=end_row_for_date, end_column=len(headers))

            # --- Final Formatting ---
            for column in ws.columns:
                ws.column_dimensions[get_column_letter(column[0].column)].width = 25
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value is not None:
                        cell.border = thin_border
            return True
        except Exception as e:
            print(f"Error creating daily tank status sheet: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    
    def _create_system_alerts_sheet(wb, results):
        """Creates a new worksheet for the System Alerts log."""
        try:
            ws = wb.create_sheet("System Alerts")
            alerts = results.get('alerts', [])

            if not alerts:
                ws.cell(row=1, column=1, value="No system alerts were generated.")
                return True

            # Define headers and styles
            headers = ['Day', 'Alert Type', 'Message']
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Define styles for different alert types
            alert_fills = {
            'danger': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'warning': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            'success': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'info': PatternFill(fill_type=None) # No background for info
            }
        
            # Populate data rows
            current_row = 2
            for alert in alerts:
                alert_type = alert.get('type', 'info')
                row_data = [
                    alert.get('day', 'N/A'),
                    alert_type.title(),
                    alert.get('message', '')
                ]
            
                fill = alert_fills.get(alert_type, alert_fills['info'])

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.fill = fill
            
                current_row += 1

            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 120
        
            return True
        except Exception as e:
           print(f"Error creating system alerts sheet: {str(e)}")
           return False
    
    def _create_simulation_data_sheet(wb, results):
        """Sheet 1: Simulation Data - Raw day-by-day simulation data"""
        try:
            ws = wb.create_sheet("Simulation Data")
            
            # Add timestamp
            timestamp_str = f"Charts Generated On: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}"
            timestamp_cell = ws.cell(row=1, column=1, value=timestamp_str)
            timestamp_cell.font = Font(bold=True, italic=True, color="4F4F4F")
            
            simulation_data = results.get('simulation_data', [])
            if not simulation_data:
                ws.cell(row=3, column=1, value="No simulation data available.")
                return True
            
            current_row = 3
            ws.cell(row=current_row, column=1, value="DAILY SIMULATION DATA").font = Font(bold=True, size=14)
            current_row += 2
            
            # Headers
            num_tanks = int(results.get('parameters', {}).get('numTanks', 12))
            headers = ['DATE', 'DAY', 'START_INVENTORY', 'PROCESSED', 'CARGO_ARRIVALS', 'END_INVENTORY']
            for i in range(1, num_tanks + 1):
                headers.append(f'TK{i}_STATUS')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            current_row += 1
            
            # Data rows
            for day_data in simulation_data:
                row_data = [
                    day_data.get('date', ''),
                    day_data.get('day', ''),
                    day_data.get('start_inventory', 0),
                    day_data.get('processing', 0),
                    day_data.get('arrivals', 0),
                    day_data.get('end_inventory', 0)
                ]
                
                # Tank statuses


                for i in range(1, num_tanks + 1):
                    status = day_data.get(f'tank{i}_status', 'N/A')
                    row_data.append(status)
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Format numbers with thousands separator
                    if isinstance(value, (int, float)) and col > 2:
                        cell.number_format = '#,##0'
                        
                current_row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            return True
            
        except Exception as e:
            print(f"Error creating simulation data sheet: {str(e)}")
            return False

    def _create_summary_analysis_sheet(wb, results):
        """Sheet 2: Summary Analysis - Key metrics and KPIs"""
        try:
            ws = wb.create_sheet("Summary Analysis")
            
            # Add timestamp
            timestamp_str = f"Charts Generated On: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}"
            timestamp_cell = ws.cell(row=1, column=1, value=timestamp_str)
            timestamp_cell.font = Font(bold=True, italic=True, color="4F4F4F")
            
            current_row = 3
            parameters = results.get('parameters', {})
            metrics = results.get('metrics', {})
            
            # Simulation Summary
            ws.cell(row=current_row, column=1, value="SIMULATION SUMMARY").font = Font(bold=True, size=14)
            current_row += 1
            ws.cell(row=current_row, column=1, value="=" * 50).font = Font(bold=True)
            current_row += 1
            
            summary_data = [
                ('Processing Rate:', f"{parameters.get('processingRate', 0):,.0f} bbl/day"),
                ('Total Days Simulated:', f"{parameters.get('schedulingWindow', 0)} days"),
                ('Tank Capacity (each):', f"{parameters.get('tankCapacity', 0):,.0f} bbl"),
                ('Processing Efficiency:', f"{metrics.get('processing_efficiency', 0):.1f}%"),
                ('Sustainable Processing:', "Yes" if metrics.get('sustainable_processing', False) else "No")
            ]
            
            for label, value in summary_data:
                ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=value)
                current_row += 1
            
            current_row += 2
            
            # Cargo Summary
            ws.cell(row=current_row, column=1, value="CARGO SUMMARY").font = Font(bold=True, size=14)
            current_row += 1
            ws.cell(row=current_row, column=1, value="=" * 50).font = Font(bold=True)
            current_row += 1
            
            cargo_data = [
                ('Total Cargoes:', metrics.get('total_cargoes', 0)),
                ('VLCC Cargoes:', f"{parameters.get('vlccCapacity', 0):,.0f} bbl capacity"),
                ('Suezmax Cargoes:', f"{parameters.get('suezmaxCapacity', 0):,.0f} bbl capacity"),
                ('Aframax Cargoes:', f"{parameters.get('aframaxCapacity', 0):,.0f} bbl capacity"),
                ('Cargo Mix:', metrics.get('cargo_mix', 'N/A'))
            ]
            
            for label, value in cargo_data:
                ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=value)
                current_row += 1
            
            current_row += 2
            
            # Inventory Summary
            ws.cell(row=current_row, column=1, value="INVENTORY SUMMARY").font = Font(bold=True, size=14)
            current_row += 1
            ws.cell(row=current_row, column=1, value="=" * 50).font = Font(bold=True)
            current_row += 1
            
            inventory_data = [
                ('Minimum Inventory Threshold:', f"{parameters.get('minInventory', 0):,.0f} bbl"),
                ('Maximum Inventory Threshold:', f"{parameters.get('maxInventory', 0):,.0f} bbl"),
                ('Minimum Reached:', f"{metrics.get('min_inventory', 0):,.0f} bbl"),
                ('Maximum Reached:', f"{metrics.get('max_inventory', 0):,.0f} bbl"),
                ('Clash Days:', metrics.get('clash_days', 0))
            ]
            
            for label, value in inventory_data:
                ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
                cell = ws.cell(row=current_row, column=2, value=value)
                
                # Color code based on values
                if "Minimum Reached" in label and isinstance(value, str) and "bbl" in value:
                    min_val = int(value.replace(',', '').replace(' bbl', ''))
                    min_threshold = parameters.get('minInventory', 0)
                    if min_val < min_threshold:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                current_row += 1
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 25
            
            return True
            
        except Exception as e:
            print(f"Error creating summary analysis sheet: {str(e)}")
            return False

    def _create_inventory_chart_sheet(wb, results):
        """Sheet 3: Inventory Chart - Line chart showing inventory levels over time"""
        try:
            ws = wb.create_sheet("Inventory Chart")
            
            simulation_data = results.get('simulation_data', [])
            parameters = results.get('parameters', {})
            
            if not simulation_data:
                ws.cell(row=1, column=1, value="No inventory data available for chart.")
                return True
            
            # Add chart data
            current_row = 1
            ws.cell(row=current_row, column=1, value="Day").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value="Inventory (bbls)").font = Font(bold=True)
            ws.cell(row=current_row, column=3, value="Min Threshold").font = Font(bold=True)
            ws.cell(row=current_row, column=4, value="Max Threshold").font = Font(bold=True)
            current_row += 1
            
            min_threshold = parameters.get('minInventory', 0)
            max_threshold = parameters.get('maxInventory', 0)
            
            for day_data in simulation_data:
                ws.cell(row=current_row, column=1, value=f"Day {day_data.get('day', 0)}")
                ws.cell(row=current_row, column=2, value=day_data.get('start_inventory', 0))
                ws.cell(row=current_row, column=3, value=min_threshold)
                ws.cell(row=current_row, column=4, value=max_threshold)
                current_row += 1
            
            # Create line chart with basic properties that work
            chart = LineChart()
            chart.title = "Daily Inventory Levels"
            chart.style = 2
            chart.y_axis.title = 'Inventory (bbls)'
            chart.x_axis.title = 'Days'
            
            # Make chart bigger
            chart.width = 20
            chart.height = 12
            
            # Data for chart
            data = Reference(ws, min_col=2, min_row=1, max_row=current_row-1, max_col=4)
            cats = Reference(ws, min_col=1, min_row=2, max_row=current_row-1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Simple line colors that definitely work
            try:
                if len(chart.series) >= 1:
                    chart.series[0].graphicalProperties.line.solidFill = "0070C0"  # Blue
                if len(chart.series) >= 2:
                    chart.series[1].graphicalProperties.line.solidFill = "FF0000"  # Red
                if len(chart.series) >= 3:
                    chart.series[2].graphicalProperties.line.solidFill = "00B050"  # Green
            except:
                # If line customization fails, continue anyway
                pass
            
            # Position chart
            ws.add_chart(chart, "F2")
            
            # Add summary statistics
            summary_row = current_row + 2
            ws.cell(row=summary_row, column=1, value="INVENTORY STATISTICS").font = Font(bold=True, size=12)
            summary_row += 1
            
            # Calculate statistics
            inventories = [day_data.get('start_inventory', 0) for day_data in simulation_data]
            if inventories:
                stats_data = [
                    ('Starting Inventory:', f"{inventories[0]:,.0f} bbls"),
                    ('Ending Inventory:', f"{inventories[-1]:,.0f} bbls"),
                    ('Maximum Inventory:', f"{max(inventories):,.0f} bbls"),
                    ('Minimum Inventory:', f"{min(inventories):,.0f} bbls"),
                    ('Average Inventory:', f"{sum(inventories)/len(inventories):,.0f} bbls"),
                    ('Min Threshold:', f"{min_threshold:,.0f} bbls"),
                    ('Max Threshold:', f"{max_threshold:,.0f} bbls")
                ]
                
                for label, value in stats_data:
                    ws.cell(row=summary_row, column=1, value=label).font = Font(bold=True)
                    ws.cell(row=summary_row, column=2, value=value)
                    summary_row += 1
            
            # Format columns
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            
            return True
            
        except Exception as e:
            print(f"Error creating inventory chart sheet: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def _create_processing_chart_sheet(wb, results):
        """Sheet 4: Processing Chart - Bar chart showing processing data"""
        try:
            ws = wb.create_sheet("Processing Chart")
            
            simulation_data = results.get('simulation_data', [])
            parameters = results.get('parameters', {})
            
            if not simulation_data:
                ws.cell(row=1, column=1, value="No processing data available for chart.")
                return True
            
            # Add chart data
            current_row = 1
            ws.cell(row=current_row, column=1, value="Day").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value="Processed").font = Font(bold=True)
            ws.cell(row=current_row, column=3, value="Target").font = Font(bold=True)
            current_row += 1
            
            target_rate = parameters.get('processingRate', 50000)
            
            for day_data in simulation_data:
                ws.cell(row=current_row, column=1, value=day_data.get('day', 0))
                ws.cell(row=current_row, column=2, value=day_data.get('processing', 0))
                ws.cell(row=current_row, column=3, value=target_rate)
                current_row += 1
            
            # Create bar chart
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Daily Processing Volumes"
            chart.y_axis.title = 'Volume (bbls)'
            chart.x_axis.title = 'Days'
            
            # Data for chart
            data = Reference(ws, min_col=2, min_row=1, max_row=current_row-1, max_col=3)
            cats = Reference(ws, min_col=1, min_row=2, max_row=current_row-1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Position chart
            ws.add_chart(chart, "E2")
            
            return True
            
        except Exception as e:
            print(f"Error creating processing chart sheet: {str(e)}")
            return False

    def _create_tank_utilization_sheet(wb, results):
        """Sheet 5: Tank Utilization - Stacked bar chart showing tank usage"""
        try:
            ws = wb.create_sheet("Tank Utilization")
            num_tanks = int(results.get('parameters', {}).get('numTanks', 12))
            
            simulation_data = results.get('simulation_data', [])
            
            if not simulation_data:
                ws.cell(row=1, column=1, value="No tank utilization data available.")
                return True
            
            # Add chart data
            current_row = 1
            headers = ['Day', 'Date','READY', 'FEEDING', 'EMPTY', 'FILLING', 'SETTLING', 'LAB TEST', 'SUSPENDED', 'FILLED']
            for col, header in enumerate(headers, 1):
                ws.cell(row=current_row, column=col, value=header).font = Font(bold=True)
            current_row += 1
            
            # Count tank statuses per day
            for day_data in simulation_data:
                # Initialize with all possible status types
                status_counts = {
                    'READY': 0, 'FEEDING': 0, 'EMPTY': 0, 'FILLING': 0, 
                    'SETTLING': 0, 'LAB TEST': 0, 'SUSPENDED': 0, 'FILLED': 0
                }
                
                for i in range(1, num_tanks + 1):
                    status = day_data.get(f'tank{i}_status', 'N/A')
                    # Handle different status formats
                    if isinstance(status, str):
                        status = status.upper().strip()
                        if status in status_counts:
                            status_counts[status] += 1
                        elif status == 'LAB_TESTING':
                            status_counts['LAB TEST'] += 1
                        elif status == 'N/A' or status == '':
                            status_counts['EMPTY'] += 1
                
                row_data = [
                    day_data.get('day', 0),
                    day_data.get('date', ''),
                    status_counts['READY'],
                    status_counts['FEEDING'],
                    status_counts['EMPTY'],
                    status_counts['FILLING'],
                    status_counts['SETTLING'],
                    status_counts['LAB TEST'],
                    status_counts['SUSPENDED'],
                    status_counts['FILLED']
                ]
                
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=current_row, column=col, value=value)
                current_row += 1
            
            # Debug: Add a summary row to see totals
            summary_row = current_row + 1
            ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
            for col in range(2, len(headers) + 1):
                formula = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{current_row-1})"
                cell = ws.cell(row=summary_row, column=col, value=formula)
                cell.font = Font(bold=True)
            
            # Create stacked bar chart
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "stacked"
            chart.overlap = 100
            chart.title = "Tank Utilization by Status"
            chart.y_axis.title = 'Number of Tanks'
            chart.x_axis.title = 'Days'
            
            # Data for chart (exclude summary row)
            data = Reference(ws, min_col=2, min_row=1, max_row=current_row-1, max_col=len(headers))
            cats = Reference(ws, min_col=1, min_row=2, max_row=current_row-1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Position chart
            ws.add_chart(chart, "K2")
            
            return True
            
        except Exception as e:
            print(f"Error creating tank utilization sheet: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def _create_cargo_arrivals_sheet(wb, results):
        """Sheet 6: Cargo Arrivals - Timeline chart of vessel movements"""
        try:
            ws = wb.create_sheet("Cargo Arrivals")
            
            cargo_report = results.get('cargo_report', [])
            
            if not cargo_report:
                ws.cell(row=1, column=1, value="No cargo arrival data available.")
                return True
            
            current_row = 1
            ws.cell(row=current_row, column=1, value="CARGO ARRIVALS TIMELINE").font = Font(bold=True, size=14)
            current_row += 2
            
            headers = ['Cargo Type', 'Arrival Date', 'Arrival Time', 'Departure Date', 'Departure Time', 'Duration (Days)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            current_row += 1
            
            for cargo in cargo_report:
                arrival_time = cargo.get('arrival_time', '')
                departure_time = cargo.get('dep_unload_port', '')
                
                arrival_date, arrival_time_only = '', ''
                if arrival_time and isinstance(arrival_time, str) and '/' in arrival_time:
                    parts = arrival_time.split(' ')
                    arrival_date = parts[0]
                    arrival_time_only = parts[1] if len(parts) > 1 else ''
                
                departure_date, departure_time_only = '', ''
                if departure_time and isinstance(departure_time, str) and '/' in departure_time:
                    parts = departure_time.split(' ')
                    departure_date = parts[0]
                    departure_time_only = parts[1] if len(parts) > 1 else ''
                
                # Calculate duration
                duration = ''
                try:
                    if arrival_time and departure_time:
                        arr_dt = _parse_json_datetime(arrival_time)
                        dep_dt = _parse_json_datetime(departure_time)
                        if arr_dt and dep_dt:
                            time_difference = dep_dt - arr_dt
                            total_seconds = time_difference.total_seconds()

                            days = int(total_seconds // 86400)
                            hours = int((total_seconds % 86400) // 3600)
                            minutes = int((total_seconds % 3600) // 60)
                            duration = f"{days}d, {hours}h, {minutes}m"
                
                except:
                    duration = 'N/A'
                
                row_data = [
                    cargo.get('type', '').title(),
                    arrival_date, arrival_time_only,
                    departure_date, departure_time_only,
                    duration
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Color code by cargo type
                    cargo_type = cargo.get('type', '').lower()
                    if 'vlcc' in cargo_type:
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    elif 'suezmax' in cargo_type:
                        cell.fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
                    elif 'aframax' in cargo_type:
                        cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                
                current_row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            return True
            
        except Exception as e:
            print(f"Error creating cargo arrivals sheet: {str(e)}")
            return False

    def _create_alerts_warnings_sheet(wb, results):
        """Sheet 7: Alerts & Warnings - Color-coded issues and warnings"""
        try:
            ws = wb.create_sheet("Alerts & Warnings")
            
            current_row = 1
            ws.cell(row=current_row, column=1, value="SYSTEM ALERTS & WARNINGS").font = Font(bold=True, size=14)
            current_row += 2
            
            headers = ['Priority', 'Alert Type', 'Description', 'Day', 'Recommended Action']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            current_row += 1
            
            # Generate alerts based on simulation data
            simulation_data = results.get('simulation_data', [])
            parameters = results.get('parameters', {})
            metrics = results.get('metrics', {})
            
            alerts = []
            
            # Check inventory thresholds
            min_threshold = parameters.get('minInventory', 0)
            max_threshold = parameters.get('maxInventory', 0)
            
            for day_data in simulation_data:
                inventory = day_data.get('start_inventory', 0)
                day = day_data.get('day', 0)
                
                if inventory < min_threshold:
                    alerts.append({
                        'priority': 'HIGH',
                        'type': 'Low Inventory',
                        'description': f'Inventory {inventory:,.0f} below minimum threshold {min_threshold:,.0f}',
                        'day': day,
                        'action': 'Schedule urgent cargo delivery'
                    })
                elif inventory > max_threshold:
                    alerts.append({
                        'priority': 'MEDIUM',
                        'type': 'High Inventory',
                        'description': f'Inventory {inventory:,.0f} above maximum threshold {max_threshold:,.0f}',
                        'day': day,
                        'action': 'Increase processing rate or defer cargo'
                    })
            
            # Check processing efficiency
            if metrics.get('processing_efficiency', 100) < 95:
                alerts.append({
                    'priority': 'MEDIUM',
                    'type': 'Processing Efficiency',
                    'description': f'Processing efficiency {metrics.get("processing_efficiency", 0):.1f}% below target',
                    'day': 'Overall',
                    'action': 'Review tank scheduling and optimize feeding sequence'
                })
            
            # Check for clash days
            if metrics.get('clash_days', 0) > 0:
                alerts.append({
                    'priority': 'HIGH',
                    'type': 'Cargo Clashes',
                    'description': f'{metrics.get("clash_days", 0)} days with cargo scheduling conflicts',
                    'day': 'Multiple',
                    'action': 'Reschedule cargo arrivals to avoid conflicts'
                })
            
            # Add sustainability warning
            if not metrics.get('sustainable_processing', True):
                alerts.append({
                    'priority': 'CRITICAL',
                    'type': 'Unsustainable Operations',
                    'description': 'Current schedule may not sustain continuous processing',
                    'day': 'Overall',
                    'action': 'Increase cargo frequency or review processing requirements'
                })
            
            # Sort alerts by priority
            priority_order = {'CRITICAL': 1, 'HIGH': 2, 'MEDIUM': 3, 'LOW': 4}
            alerts.sort(key=lambda x: priority_order.get(x['priority'], 5))
            
            if not alerts:
                ws.cell(row=current_row, column=1, value="No alerts or warnings detected.").font = Font(color="00B050")
                current_row += 1
            else:
                for alert in alerts:
                    row_data = [
                        alert['priority'],
                        alert['type'],
                        alert['description'],
                        alert['day'],
                        alert['action']
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.alignment = Alignment(horizontal='center' if col == 1 else 'left', wrap_text=True)
                        
                        # Color code by priority
                        if alert['priority'] == 'CRITICAL':
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)
                        elif alert['priority'] == 'HIGH':
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif alert['priority'] == 'MEDIUM':
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    
                    current_row += 1
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 40
            
            return True
            
        except Exception as e:
            print(f"Error creating alerts warnings sheet: {str(e)}")
            return False

    def _create_cargo_schedule_sheet(wb, results):
        """Sheet 8: Cargo Schedule - Detailed cargo scheduling information"""
        try:
            ws = wb.create_sheet("Cargo Schedule")
            
            cargo_report = results.get('cargo_report', [])
            parameters = results.get('parameters', {})
            
            current_row = 1
            ws.cell(row=current_row, column=1, value="DETAILED CARGO SCHEDULE").font = Font(bold=True, size=14)
            current_row += 2
            
            # Schedule parameters
            ws.cell(row=current_row, column=1, value="SCHEDULE PARAMETERS").font = Font(bold=True, size=12)
            current_row += 1
            
            params_data = [
                ('Pre-Journey Time:', f"{parameters.get('preJourneyDays', 0)} days"),
                ('Journey Time:', f"{parameters.get('journeyDays', 0)} days"),
                ('Pre-Discharge Time:', f"{parameters.get('preDischargeDays', 0)} days"),
                ('Settling Time:', f"{parameters.get('settlingTime', 0)} days"),
                ('Lab Testing Time:', f"{parameters.get('labTestingDays', 0)} days"),
                ('Buffer Days:', f"{parameters.get('bufferDays', 0)} days"),
                ('Pumping Rate:', f"{parameters.get('pumpingRate', 0):,.0f} bbl/hr")
            ]
            
            for label, value in params_data:
                ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=value)
                current_row += 1
            
            current_row += 2
            
            # Detailed cargo information
            ws.cell(row=current_row, column=1, value="CARGO DETAILS").font = Font(bold=True, size=12)
            current_row += 1
            
            headers = ['Cargo','BERTH','Type', 'Size (bbls)', 'Departure', 'Arrival', 'Discharge', 'Pumping Hrs', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            current_row += 1
            
            if not cargo_report:
                ws.cell(row=current_row, column=1, value="No cargo schedule data available.")
                return True
            
            cargo_counter = 1
            for cargo in cargo_report:
                try:
                    vessel_name = str(cargo.get('vessel_name', 'Unknown')).title()
                    cargo_size = cargo.get('size', 0)
                    
                    # Safely get cargo size as number
                    if isinstance(cargo_size, str):
                        cargo_size = float(cargo_size.replace(',', '')) if cargo_size.replace(',', '').replace('.', '').isdigit() else 0
                    elif not isinstance(cargo_size, (int, float)):
                        cargo_size = 0
                    
                    # Calculate pumping hours safely
                    pumping_rate = parameters.get('pumpingRate', 30000)
                    if pumping_rate and pumping_rate > 0:
                        pumping_hours = cargo_size / pumping_rate
                    else:
                        pumping_hours = 0
                    
                    # Safely get time strings
                    departure_time = str(cargo.get('dep_time', 'N/A'))
                    arrival_time = str(cargo.get('arrival_time', 'N/A'))
                    discharge_time = str(cargo.get('dep_unload_port', 'N/A'))
                    
                    row_data = [
                        f"Cargo {cargo_counter}",
                        cargo.get('berth', 'N/A'),  # <-- ADD THIS LINE
                        vessel_name,
                        cargo_size,  # Will be formatted by Excel
                        departure_time,
                        arrival_time,
                        discharge_time,
                        f"{pumping_hours:.1f}",
                        cargo.get('status', 'Scheduled') # <-- THIS IS THE CHANGE
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.alignment = Alignment(horizontal='center')
                        
                        # Format cargo size with thousands separator
                        if col == 3 and isinstance(cargo_size, (int, float)) and cargo_size > 0:
                            cell.number_format = '#,##0'
                    
                    current_row += 1
                    cargo_counter += 1
                    
                except Exception as cargo_error:
                    print(f"Error processing cargo {cargo_counter}: {str(cargo_error)}")
                    # Continue with next cargo instead of failing entire sheet
                    cargo_counter += 1
                    continue
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            return True
            
        except Exception as e:
            print(f"Error creating cargo schedule sheet: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def _create_cargo_timeline_sheet(wb, results):
        """Sheet 9: Cargo Timeline - Visual timeline showing cargo movements with sizes"""
        try:
            ws = wb.create_sheet("Cargo Timeline")
            
            cargo_report = results.get('cargo_report', [])
            parameters = results.get('parameters', {})
            
            if not cargo_report:
                ws.cell(row=1, column=1, value="No cargo timeline data available.")
                return True
            
            current_row = 1
            ws.cell(row=current_row, column=1, value="CARGO MOVEMENT TIMELINE").font = Font(bold=True, size=14)
            current_row += 2
            
            # Define cargo type properties (largest to smallest)
            cargo_types = {
                'VLCC': {'size': 2000000, 'color': '1f4e79', 'height': 5, 'priority': 1},
                'SUEZMAX': {'size': 1000000, 'color': '2e75b6', 'height': 4, 'priority': 2},
                'AFRAMAX': {'size': 750000, 'color': '5b9bd5', 'height': 3, 'priority': 3},
                'PANAMAX': {'size': 600000, 'color': '9cc3e5', 'height': 2, 'priority': 4},
                'HANDYMAX': {'size': 350000, 'color': 'c5dbef', 'height': 1, 'priority': 5}
            }
            
            # Create timeline data structure
            timeline_data = []
            
            for i, cargo in enumerate(cargo_report):
                cargo_type = str(cargo.get('type', 'UNKNOWN')).upper()
                cargo_size = cargo.get('size', 0)
                
                # Try to parse dates
                try:
                    arrival_str = cargo.get('arrival_time', '')
                    departure_str = cargo.get('dep_unload_port', '')
                    
                    if arrival_str and departure_str:
                        arrival_dt = _parse_json_datetime(arrival_str)
                        departure_dt = _parse_json_datetime(departure_str)
                        
                        if arrival_dt and departure_dt:
                            duration_days = (departure_dt - arrival_dt).days
                            
                            timeline_data.append({
                                'cargo_num': i + 1,
                                'type': cargo_type,
                                'size': cargo_size,
                                'arrival': arrival_dt,
                                'departure': departure_dt,
                                'duration': duration_days,
                                'arrival_day': arrival_str.split(' ')[0] if ' ' in arrival_str else arrival_str,
                                'departure_day': departure_str.split(' ')[0] if ' ' in departure_str else departure_str,
                                'vessel_name': cargo.get('vessel_name', f"Cargo {i+1}")
                            })
                except Exception as e:
                    print(f"Error parsing dates for cargo {i+1}: {e}")
                    continue
            
            if not timeline_data:
                ws.cell(row=current_row, column=1, value="Unable to parse cargo timeline data.")
                return True
            
            # Sort by arrival time
            timeline_data.sort(key=lambda x: x['arrival'])
            
            # Create headers for timeline table
            headers = ['Vessel Type', 'Type', 'Size (bbls)', 'Arrival', 'Departure', 'Duration', 'Visual Timeline']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            current_row += 1
            
            # Create visual timeline
            max_duration = max([cargo['duration'] for cargo in timeline_data]) if timeline_data else 10
            timeline_start_col = 8  # Column H
            
            for cargo in timeline_data:
                cargo_type = cargo['type']
                type_info = cargo_types.get(cargo_type, cargo_types['HANDYMAX'])  # Default to smallest
                
                # Basic cargo info
                ws.cell(row=current_row, column=1, value=cargo.get('vessel_name', '')).alignment = Alignment(horizontal='center')
                ws.cell(row=current_row, column=2, value=cargo_type).fill = PatternFill(
                    start_color=type_info['color'], 
                    end_color=type_info['color'], 
                    fill_type="solid"
                )
                
                size_cell = ws.cell(row=current_row, column=3, value=cargo['size'])
                size_cell.number_format = '#,##0'
                
                ws.cell(row=current_row, column=4, value=cargo['arrival_day'])
                ws.cell(row=current_row, column=5, value=cargo['departure_day'])
                ws.cell(row=current_row, column=6, value=f"{cargo['duration']} days")
                
                # Create visual timeline bars - different widths based on cargo size
                bar_width = min(max(cargo['duration'], 2), 15)  # Minimum 2, maximum 15 columns
                bar_height = type_info['height']
                
                # Fill cells to represent cargo movement timeline
                for day_offset in range(bar_width):
                    col = timeline_start_col + day_offset
                    
                    # Create multiple rows for height effect
                    for height_row in range(bar_height):
                        cell = ws.cell(row=current_row + height_row, column=col, value="")
                        
                        # Color intensity based on cargo size (darker for larger)
                        if cargo_type == 'VLCC':
                            cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
                        elif cargo_type == 'SUEZMAX':
                            cell.fill = PatternFill(start_color="2e75b6", end_color="2e75b6", fill_type="solid")
                        elif cargo_type == 'AFRAMAX':
                            cell.fill = PatternFill(start_color="5b9bd5", end_color="5b9bd5", fill_type="solid")
                        elif cargo_type == 'PANAMAX':
                            cell.fill = PatternFill(start_color="9cc3e5", end_color="9cc3e5", fill_type="solid")
                        else:  # HANDYMAX
                            cell.fill = PatternFill(start_color="c5dbef", end_color="c5dbef", fill_type="solid")
                        
                        # Add border for definition
                        cell.border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )
                
                # Move to next cargo (accounting for height)
                current_row += max(bar_height, 1) + 1
            
            # Add legend
            legend_row = current_row + 2
            ws.cell(row=legend_row, column=1, value="LEGEND - Cargo Types (Largest to Smallest)").font = Font(bold=True, size=12)
            legend_row += 1
            
            for cargo_type, info in cargo_types.items():
                ws.cell(row=legend_row, column=1, value=cargo_type).font = Font(bold=True)
                legend_cell = ws.cell(row=legend_row, column=2, value=f"{info['size']:,} bbls capacity")
                legend_cell.fill = PatternFill(start_color=info['color'], end_color=info['color'], fill_type="solid")
                legend_cell.font = Font(color="FFFFFF" if info['priority'] <= 2 else "000000")
                legend_row += 1
            
            # Add timeline scale
            scale_row = 3
            ws.cell(row=scale_row, column=timeline_start_col - 1, value="Timeline →").font = Font(bold=True, italic=True)
            for day in range(1, 21):  # Show first 20 days
                ws.cell(row=scale_row, column=timeline_start_col + day - 1, value=f"D{day}").font = Font(size=8)
            
            # Auto-adjust columns
            for col in range(1, timeline_start_col):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Make timeline columns narrower for better visual effect
            for col in range(timeline_start_col, timeline_start_col + 20):
                ws.column_dimensions[get_column_letter(col)].width = 3
            
            return True
            
        except Exception as e:
            print(f"Error creating cargo timeline sheet: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    @app.route('/api/calculate_buffer_stock', methods=['POST'])
    def calculate_buffer_stock():
        """Calculate buffer stock for continuous operation"""
        try:
            params = request.json
            buffer_info = scheduler._calculate_buffer_stock(params)
            return jsonify({
                'success': True,
                'buffer_info': buffer_info
            })
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400

    @app.route('/api/calculate_pumping_days', methods=['POST'])
    def calculate_pumping_days():
        """Calculate pumping days for given cargo size"""
        try:
            params = request.json
            cargo_size = float(params['cargoSize']) 
            pumping_rate = float(params.get('pumpingRate', 30000))
            
            pumping_days = cargo_size / (pumping_rate * 24) if pumping_rate > 0 else 0
            
            return jsonify({
                'success': True,
                'pumping_days': round(pumping_days, 2)
            })
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400

    @app.route('/api/export_tank_status', methods=['POST'])
    def export_tank_status():
        """Export sequence report with both Sequence Summary and Tank Filling Volumes sheets"""
        try:
            results = request.json
            
            # Create a new workbook
            wb = Workbook()
            
            # Remove the default sheet since we'll create our own
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create the sequence summary sheet
            sequence_success = _create_sequence_summary_sheets(wb, results)
            
            # Create the new tank filling volumes sheet
            volume_success = _create_tank_filling_volumes_sheet(wb, results)

            status_success = _create_daily_tank_status_sheet(wb, results)
            
            if not sequence_success:
                return jsonify({'error': 'Failed to create sequence summary'}), 400
                
            if not volume_success:
                return jsonify({'error': 'Failed to create tank filling volumes sheet'}), 400

            if not status_success:
                return jsonify({'error': 'Failed to create status report'}), 400
            
            # Create a proper temporary file that gets deleted immediately after sending
            import tempfile
            import os
            
            # Generate download filename with timestamp
            timestamp_str = datetime.now().strftime('%d-%b-%Y_%H-%M-%S')
            download_filename = f"sequence_report_{timestamp_str}.xlsx"
            
            # Create temporary file with context manager for auto-cleanup
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                wb.save(tmp_file.name)
                tmp_file_path = tmp_file.name
            
            def remove_file():
                try:
                    os.unlink(tmp_file_path)
                except:
                    pass
            
            # Send file and schedule cleanup
            response = send_file(
                tmp_file_path,
                as_attachment=True,
                download_name=download_filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheet.sheet'
            )
            
            # Clean up temp file immediately
            remove_file()
            
            return response
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Sequence report export failed: {str(e)}'}), 400

    @app.route('/api/export_charts', methods=['POST'])
    def export_charts():
        """Export comprehensive charts workbook with 9 sheets including embedded charts and cargo timeline"""
        try:
            results = request.json
            
            # Create workbook with timestamp in workbook name
            wb = Workbook()
            timestamp_str = datetime.now().strftime('%d-%b-%Y %H:%M:%S')
            wb.title = f"charts {timestamp_str}"
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create all 9 sheets
            success_results = {}
            success_results['sheet1'] = _create_simulation_data_sheet(wb, results)
            success_results['sheet2'] = _create_summary_analysis_sheet(wb, results)
            success_results['sheet3'] = _create_inventory_chart_sheet(wb, results)
            success_results['sheet4'] = _create_processing_chart_sheet(wb, results)
            success_results['sheet5'] = _create_tank_utilization_sheet(wb, results)
            success_results['sheet6'] = _create_cargo_arrivals_sheet(wb, results)
            success_results['sheet7'] = _create_alerts_warnings_sheet(wb, results)
            success_results['sheet8'] = _create_cargo_schedule_sheet(wb, results)
            success_results['sheet9'] = _create_cargo_timeline_sheet(wb, results)
            success_results['sheet9'] = _create_system_alerts_sheet(wb, results) # <-- ADD THIS LINE
            
            # Check if any sheet creation failed
            failed_sheets = [k for k, v in success_results.items() if not v]
            if failed_sheets:
                return jsonify({'error': f'Failed to create sheets: {", ".join(failed_sheets)}'}), 400
            
            # Create a proper temporary file that gets deleted immediately after sending
            import tempfile
            import os
            
            # Generate download filename with timestamp
            timestamp_str_file = datetime.now().strftime('%d-%b-%Y_%H-%M-%S')
            download_filename = f"charts_report_{timestamp_str_file}.xlsx"
            
            # Create temporary file with context manager for auto-cleanup
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                wb.save(tmp_file.name)
                tmp_file_path = tmp_file.name
            
            def remove_file():
                try:
                    os.unlink(tmp_file_path)
                except:
                    pass
            
            # Send file and schedule cleanup
            response = send_file(
                tmp_file_path,
                as_attachment=True,
                download_name=download_filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheet.sheet'
            )
            
            # Clean up temp file immediately
            remove_file()
            
            return response
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Charts export failed: {str(e)}'}), 400

    @app.route('/api/buffer_analysis', methods=['POST'])
    def buffer_analysis():
        try:
            params = request.json
            num_tanks = int(params.get('numTanks', 12))
            processing_rate = float(params.get('processingRate', 50000))
            tank_capacity = float(params.get('tankCapacity', 500000))
            journey_days = int(params.get('journeyDays', 10))
            buffer_days = int(params.get('bufferDays', 2))
            pre_journey_days = float(params.get('preJourneyDays', 1))
            pre_discharge_days = float(params.get('preDischargeDays', 1))
            settling_days = float(params.get('settlingTime', 2))
            lab_testing_days = float(params.get('labTestingDays', 1))
            
            buffer_results = {}
            
            # Normal operations buffer
            normal_lead_time = pre_journey_days + journey_days + pre_discharge_days + settling_days + lab_testing_days
            normal_buffer_days = normal_lead_time + buffer_days + 3
            normal_buffer_needed = processing_rate * normal_buffer_days
            normal_tanks_needed = max(1, int(normal_buffer_needed / tank_capacity) + 1) if tank_capacity > 0 else 1
            
            

            buffer_results['normal_operations'] = {
                'description': 'Normal Operations Buffer',
                'lead_time': normal_lead_time,
                'buffer_needed': normal_buffer_needed,
                'tanks_needed': normal_tanks_needed,
                'adequate_current': normal_tanks_needed <= num_tanks,
                'additional_tanks': max(0, normal_tanks_needed - num_tanks)
            }
            
            # Extended disruption buffer
            disruption_duration = int(params.get('disruptionDuration', 7))
            disruption_buffer_days = normal_lead_time + buffer_days + disruption_duration
            disruption_buffer_needed = processing_rate * disruption_buffer_days
            disruption_tanks_needed = max(1, int(disruption_buffer_needed / tank_capacity) + 1) if tank_capacity > 0 else 1
            
            buffer_results['extended_disruption'] = {
                'description': f'Extended Disruption Buffer ({disruption_duration} days)',
                'lead_time': normal_lead_time,
                'buffer_needed': disruption_buffer_needed,
                'tanks_needed': disruption_tanks_needed,
                'adequate_current': disruption_tanks_needed <= num_tanks,
                'additional_tanks': max(0, disruption_tanks_needed - num_tanks)
            }
            
            # Emergency scenario buffer
            emergency_buffer_days = normal_lead_time + buffer_days + 14
            emergency_buffer_needed = processing_rate * emergency_buffer_days
            emergency_tanks_needed = max(1, int(emergency_buffer_needed / tank_capacity) + 1) if tank_capacity > 0 else 1
            
            buffer_results['emergency_scenario'] = {
                'description': 'Emergency Scenario Buffer (14 days)',
                'lead_time': normal_lead_time,
                'buffer_needed': emergency_buffer_needed,
                'tanks_needed': emergency_tanks_needed,
                'adequate_current': emergency_tanks_needed <= num_tanks,
                'additional_tanks': max(0, emergency_tanks_needed - num_tanks)
            }
            
            return jsonify(buffer_results)
            
        except Exception as e:
            return jsonify({'error': str(e)}), 400

    @app.route('/api/cargo_optimization', methods=['POST'])
    def cargo_optimization():
        try:
            params = request.json
            
            cargo_types = []
            if params.get('vlccCapacity', 0) > 0:
                cargo_types.append({'name': 'vlcc', 'size': float(params.get('vlccCapacity'))})
            if params.get('suezmaxCapacity', 0) > 0:
                cargo_types.append({'name': 'suezmax', 'size': float(params.get('suezmaxCapacity'))})
            if params.get('aframaxCapacity', 0) > 0:
                cargo_types.append({'name': 'aframax', 'size': float(params.get('aframaxCapacity'))})
            if params.get('panamaxCapacity', 0) > 0:
                cargo_types.append({'name': 'panamax', 'size': float(params.get('panamaxCapacity'))})
            if params.get('handymaxCapacity', 0) > 0:
                cargo_types.append({'name': 'handymax', 'size': float(params.get('handymaxCapacity'))})
            
            optimization_results = {}
            combo_counter = 1
            
            # Test single cargo types
            for cargo in cargo_types:
                test_params = params.copy()
                
                # Disable all other cargo types
                for other_cargo in ['vlcc', 'suezmax', 'aframax', 'panamax', 'handymax']:
                    if other_cargo != cargo['name']:
                        test_params[f'{other_cargo}Capacity'] = 0
                
                results = scheduler.run_simulation(test_params)
                
                if 'error' not in results:
                    metrics = results.get('metrics', {})
                    
                    optimization_results[f'combo_{combo_counter}'] = {
                        'cargo_types': [cargo['name']],
                        'efficiency': metrics.get('processing_efficiency', 0),
                        'total_cargoes': metrics.get('total_cargoes', 0),
                        'cargo_mix': metrics.get('cargo_mix', ''),
                        'clash_days': metrics.get('clash_days', 0),
                        'sustainable': metrics.get('sustainable_processing', False),
                        'min_inventory': metrics.get('min_inventory', 0)
                    }
                    combo_counter += 1
            
            # Test cargo type combinations
            if len(cargo_types) >= 2:
                for i in range(len(cargo_types)):
                    for j in range(i + 1, len(cargo_types)):
                        cargo1 = cargo_types[i]
                        cargo2 = cargo_types[j]
                        
                        test_params = params.copy()
                        for cargo_name in ['vlcc', 'suezmax', 'aframax', 'panamax', 'handymax']:
                            if cargo_name not in [cargo1['name'], cargo2['name']]:
                                test_params[f'{cargo_name}Capacity'] = 0
                        
                        results = scheduler.run_simulation(test_params)
                        
                        if 'error' not in results:
                            metrics = results.get('metrics', {})
                            
                            optimization_results[f'combo_{combo_counter}'] = {
                                'cargo_types': [cargo1['name'], cargo2['name']],
                                'efficiency': metrics.get('processing_efficiency', 0),
                                'total_cargoes': metrics.get('total_cargoes', 0),
                                'cargo_mix': metrics.get('cargo_mix', ''),
                                'clash_days': metrics.get('clash_days', 0),
                                'sustainable': metrics.get('sustainable_processing', False),
                                'min_inventory': metrics.get('min_inventory', 0)
                            }
                            combo_counter += 1
            
            return jsonify(optimization_results)
            
        except Exception as e:
            return jsonify({'error': str(e)}), 400

    @app.route('/api/timestamp_consumption_analysis', methods=['POST'])
    def timestamp_consumption_analysis():
        """Analyze timestamp-based consumption calculations for tanks"""
        try:
            params = request.json
            
            crude_processing_date = params.get('crudeProcessingDate', '')
            if crude_processing_date and ':' not in crude_processing_date:
                print("API INFO: Timestamp missing, but simulation will use a default.")
                
            results = scheduler.run_simulation(params)
            
            if 'error' in results:
                return jsonify({'success': False, 'error': results['error']}), 400
            
            timestamp_summary = scheduler._calculate_timestamp_consumption_summary(params)
            
            analysis = {
                'success': True,
                'timestamp_consumption_summary': timestamp_summary,
                'calculation_examples': [],
                'key_insights': []
            }
            
            if timestamp_summary.get('tank_consumption_details'):
                for tank_detail in timestamp_summary['tank_consumption_details'][:3]:
                    if tank_detail['daily_breakdown']:
                        example = tank_detail['daily_breakdown'][0]
                        analysis['calculation_examples'].append({
                            'tank_id': tank_detail['tank_id'],
                            'example_day': example['day'],
                            'start_time': example['start_time'],
                            'end_time': example['end_time'],
                            'consumption': example['consumption'],
                            'calculation': example['calculation_note']
                        })
            
            analysis['key_insights'] = [
                f"Processing rate: {params.get('processingRate', 50000):,.0f} bbl/day",
                f"Total tanks with feeding activity: {len(timestamp_summary.get('tank_consumption_details', []))}",
                "Consumption calculated using: (end_time - start_time) / 24 * processing_rate",
                "Next day start time always set to 00:00 hrs for continuous operation",
                "SUSPENDED tanks: Consumption = 0, timing calculated separately"
            ]
            
            return jsonify(analysis)
            
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/api/download_template', methods=['GET'])
    def download_template():
        """Download Excel template for batch simulation input"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Enhanced Simulation Template"
            
            headers = [
                'Scenario_Name', 'Processing_Rate', 'Tank_Capacity', 'Pumping_Rate',
                'Settling_Time', 'Pre_Journey_Days', 'Journey_Days', 'Pre_Discharge_Days',
                'Lab_Testing_Days', 'Buffer_Days', 'Scheduling_Window', 'crudeProcessingDate',
                'VLCC_Capacity', 'Suezmax_Capacity', 'Aframax_Capacity',
                'Panamax_Capacity', 'Handymax_Capacity'
            ]
            num_tanks = int(request.args.get('num_tanks', 12)) 
            for i in range(1, num_tanks + 1):
                headers.extend([f'Tank{i}_Level', f'DeadBottom{i}'])
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Add sample data row
            sample_row = [
                'Enhanced_Scenario_1', 50000, 500000, 30000, 2, 1, 10, 1, 1, 2, 30, '2024-01-01 08:00',
                2000000, 1000000, 750000, 400000, 300000
            ]
            
            for i in range(num_tanks):
                sample_row.extend([400000, 10000])
            
            for col, value in enumerate(sample_row, 1):
                ws.cell(row=2, column=col, value=value)
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column].width = adjusted_width
            
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            filepath = temp_file.name
            temp_file.close()
            
            wb.save(filepath)
            
            return send_file(
                filepath,
                as_attachment=True,
                download_name="enhanced_refinery_simulation_template.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheet.sheet'
            )
            
        except Exception as e:
            return jsonify({'error': f'Template generation failed: {str(e)}'}), 400

    @app.route('/api/save_inputs', methods=['POST'])
    def save_inputs():
        """Saves the user's current input parameters to a JSON file."""
        try:
            inputs = request.json
            with open(INPUTS_FILE, 'w') as f:
                json.dump(inputs, f, indent=4)
            return jsonify({'success': True, 'message': 'Inputs saved successfully.'})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/api/load_inputs', methods=['GET'])
    def load_inputs():
        """Loads the last saved input parameters from the JSON file."""
        try:
            if os.path.exists(INPUTS_FILE):
                with open(INPUTS_FILE, 'r') as f:
                    inputs = json.load(f)
                return jsonify(inputs)
            else:
                return jsonify({})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500